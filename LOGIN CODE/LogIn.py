#=================================================IMPORTED CUSTOMTKINTER THEME======================================================================
#To Instal Customtkinter Type this in cmd / visual studio code terminal - pip install customtkinter
#Also The Excel Use in each def functions is accounts.xlsx 
import customtkinter as ctk
from tkinter import messagebox
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

root = ctk.CTk()
root.geometry('600x600')
root.title('LogIN')
root._set_appearance_mode('dark')
root.resizable(False,False)

excel_con = Workbook()

#=================================================REGISTER FUNCTION======================================================================
def registerfunction():
    reg = ctk.CTkToplevel()
    reg.geometry('600x400')
    reg.resizable(False,False)
    reg.configure(bg='dimgray')
    reg.title('REGISTER')
    root.withdraw()

#=================================================ACCOUNTS(EXCEL FOR THE REGISTERED ACCOUNTS)======================================================================
    excel_con = load_workbook('accounts.xlsx')
    excel_activate = excel_con.active

    def register():
        Found = False
        user = username_ntr.get()
        password = password_ntr.get()
        email = em_ntr.get()
        
        if user == "" and password == "" and email == "":
            messagebox.showinfo("Notification", "All fields are required", parent=root)
        elif user == "" or password == "" or email == "":
            messagebox.showinfo("Notification", "All fields are required", parent=root)
        elif len(password) < 8:
            messagebox.showerror("Notification", "Password must be at least 8 characters long!", parent=root)
        else:
            for each_cell in range(2, excel_activate.max_row + 1):
                if user == excel_activate['A' + str(each_cell)].value:
                    Found = True
                    break
                else:
                    Found = False
            if Found:
                messagebox.showerror("ERROR", "Account Already Created!!!")
            else:
                lastrow = str(excel_activate.max_row + 1)
                excel_activate['A' + lastrow] = user
                excel_activate['B' + lastrow] = password
                excel_activate['C' + lastrow] = em_ntr.get()
                excel_con.save('accounts.xlsx')
                reg.destroy()
                messagebox.showinfo("SUCCESS", "Account Successfully Created!!!")
                root.deiconify()

    def go_back_to_main(root, reg):
        reg.destroy()
        root.deiconify()
        messagebox.showinfo("Login", "Back To LogIn")

    reminder_label = ctk.CTkLabel(reg,
                                  text='Put Your Email To Retrieve Or Change Your Password', 
                                  width=50,
                                  font=('arial',20,'underline'))
    user_label = ctk.CTkLabel(reg, 
                              text="USERNAME:", 
                              width=10,
                              font=('arial',20))
    pass_label = ctk.CTkLabel(reg, 
                              text="PASSWORD:", 
                              width=10,
                              font=('arial',20))
    username_ntr = ctk.CTkEntry(reg, 
                                width=200,
                                font=('arial',20),
                                placeholder_text='•USERNAME')
    password_ntr = ctk.CTkEntry(reg, 
                                width=200, 
                                show="•",
                                font=('arial',20),
                                placeholder_text='•PASSWORD')
    em_label = ctk.CTkLabel(reg, 
                            text="EMAIL:", 
                            width=10,
                            font=('arial',20))
    em_ntr = ctk.CTkEntry(reg, 
                          width=200,
                          font=('arial',20),
                          placeholder_text='•EMAIL')
    registers = ctk.CTkButton(reg,
                            text="Register",
                            command=lambda: register(),
                            width=16,
                            font=('arial',30),
                            fg_color='#323232',
                            hover_color='black')

    user_label.place(x=120, 
                    y=100)
    username_ntr.place(x=270, 
                       y=100)
    pass_label.place(x=120, 
                    y=150)
    password_ntr.place(x=270, 
                       y=150)
    em_label.place(x=120,
                   y=200)
    em_ntr.place(x=270,
                 y=200)
    registers.place(x=230,
                    y=250)
    reminder_label.place(x=50,
                        y=320)

    reg.protocol("WM_DELETE_WINDOW", lambda: go_back_to_main(root, reg))
    reg.mainloop()

#=================================================LOG-IN FUNCTION======================================================================
def loginfunction():
#=================================================ACCOUNTS(EXCEL FOR THE REGISTERED ACCOUNTS)======================================================================
    excel_con = load_workbook('accounts.xlsx')
    excel_activate = excel_con.active
    user = username.get()
    passw = password.get()
    found = False

    if user == "" or passw == "":
        messagebox.showinfo("Notification", "All fields are required")
    elif len(passw) < 8:
        messagebox.showerror("Notification", "Wrong Password / Username!\nTry Again!")
    else:
        for each_cell in range(2, excel_activate.max_row + 1):
            if (username.get() == excel_activate['A' + str(each_cell)].value and password.get() == excel_activate['B' + str(each_cell)].value):
                found = True
                break
#=================================================SHOP INTERFACE======================================================================
        if found:
            messagebox.showinfo("Notification", "LogIn Successfully")
        else:
            messagebox.showerror("Notification", "Wrong Password / Username!\n Try Again!", parent=root)

#=================================================EXIT FUNCTION======================================================================
def cancelfunction():
    root.destroy()

#=================================================CHANGE PASSWORD FUNCTION======================================================================
def change_password():
    put = ctk.CTkToplevel()
    put.geometry('400x200')
    put.title('CHANGE PASSWORD')
    put.resizable(False,False)
    put.configure(bg='dimgray')
    root.withdraw()

#=================================================ACCOUNTS(EXCEL FOR THE REGISTERED ACCOUNTS)======================================================================
    excel_con = load_workbook('accounts.xlsx')
    excel_activate = excel_con.active

#=================================================TO GO BACK TO LOGIN INTERFACE======================================================================
    def go_back_to_main(root, put):
        put.destroy()
        root.deiconify()
        messagebox.showinfo("Login", "Back To LogIn")
    
#=================================================NEW INTERFACE FOR CHANGING PASSWORD)======================================================================
    def go_to_change():
        change = put_entry.get()
        if change == "" :
            messagebox.showinfo("Notification", "Email required", parent=root)
        else:
            for each_cell in range(2, (excel_activate.max_row)+1):
                if (change == excel_activate['C'+str(each_cell)].value):
                    Found = True
                    break
                else:
                    Found = False
            if (Found == True):
                change = ctk.CTkToplevel()
                change.geometry('600x310')
                change.title('CHANGE PASSWORD')
                change.configure(bg='dimgray')
                put.withdraw()

                def update():
                    passcode = pas_ntr.get()

                    if passcode == "":
                        messagebox.showinfo("Notification", "All fields are required", parent=root)
                    elif len(passcode) < 8:
                        messagebox.showerror("Notification", "Password must be at least 8 characters long!", parent=root)
                    else:
                        excel_activate['B'+str(each_cell)].value = pas_ntr.get()
                        excel_con.save('accounts.xlsx')
                        change.destroy()
                        messagebox.showinfo("Notification","Password Changed Successfully!!!")
                        root.deiconify()
                    
                edit_label = ctk.CTkLabel(change,
                                        text='Input New Password',
                                        font=('arial',30,'underline'))
                edit_label.pack(pady=10)

                pas_label = ctk.CTkLabel(change,
                                        text="New Password :",
                                        pady=20,
                                        font=('arial',20))
                pas_ntr = ctk.CTkEntry(change,
                                    width=200,
                                    show='•',
                                    font=('arial',20),
                                    placeholder_text='•PASSWORD')
                
                pas_label.pack(pady=10)
                pas_ntr.pack(pady=10)

                updatebtn = ctk.CTkButton(change,
                                        width=15,
                                        text='Change Password',
                                        command=lambda:update(),
                                        font=('arial',30),
                                        corner_radius=10,
                                        fg_color='#323232',
                                        hover_color='black')
                updatebtn.pack(pady=10)

                change.mainloop()
            else:
                messagebox.showerror("Notification", "Wrong Email", parent=root)

#=================================================LABEL,ENTRY,BUTTON======================================================================
    put_email = ctk.CTkLabel(put,text='InPut Your Email :',
                            width=20,
                            font=('arial',20,'underline'))
    put_entry = ctk.CTkEntry(put,
                            width=200,
                            font=('arial',20),
                            placeholder_text='•EMAIL')

    put_email.pack(pady=10)
    put_entry.pack(pady=10)

    enter_btn = ctk.CTkButton(put,
                            text="Enter",
                            width=14,command=lambda:go_to_change(),
                            font=('arial',30),
                            fg_color='#323232')
    enter_btn.pack(pady=10)

#=================================================WHENEVER CLOSES IT GOES BACK TO LOGIN INTERFACE======================================================================
    put.protocol("WM_DELETE_WINDOW", lambda: go_back_to_main(root, put))
    put.mainloop()

#=================================================FORGOT PASSWORD FUNCTION======================================================================
def forgot_password():
    forgot = ctk.CTkToplevel()
    forgot.geometry('200x200')
    forgot.resizable(False,False)
    forgot.title('FORGOT PASSWORD')
    forgot.configure(bg='dimgray')
    root.withdraw()

#=================================================ACCOUNTS(EXCEL FOR THE REGISTERED ACCOUNTS)======================================================================
    excel_con = load_workbook('accounts.xlsx')
    excel_activate = excel_con.active

#=================================================TO GO BACK TO LOGIN INTERFACE======================================================================
    def go_back_to_main(root, forgot):
        forgot.destroy()
        root.deiconify()
        messagebox.showinfo("Login", "Back To LogIn")

#=================================================NEW INTERFACE TO CHANGE THE PASSWORD======================================================================
    def go_to_change():
        forgotn = forgot_entry_name.get()
        forgote = forgot_entry.get()
        each_cell = excel_activate.max_row
        if forgotn == "" or forgote == "":
            messagebox.showinfo("Notification", "All fields are required", parent=root)
        elif forgotn == "" and (forgote == excel_activate['C'+str(each_cell)].value):
            messagebox.showinfo("Notification", "All fields are required", parent=root)
        elif (forgotn == excel_activate['A'+str(each_cell)].value) and forgote == "":
            messagebox.showinfo("Notification", "All fields are required", parent=root)
        else:
            for each_cell in range(2, (excel_activate.max_row)+1):
                if (forgot_entry_name.get() == excel_activate['A'+str(each_cell)].value and forgot_entry.get() == excel_activate['C'+str(each_cell)].value):
                    Found = True
                    break
                else:
                    Found = False
            if (Found == True):
                forgot_change = ctk.CTkToplevel()
                forgot_change.geometry('600x210')
                forgot_change.title('FORGOT PASSWORD')
                forgot_change.configure(bg='dimgray')
                forgot_change.resizable(False,False)
                forgot.withdraw()

                def update():
                    passcode = pas_ntr.get()

                    if passcode == "":
                        messagebox.showinfo("Notification", "All fields are required", parent=root)
                    elif len(passcode) < 8:
                        messagebox.showerror("Notification", "Password must be at least 8 characters long!", parent=root)
                    else:
                        excel_activate['B'+str(each_cell)].value = pas_ntr.get()
                        excel_con.save('accounts.xlsx')
                        forgot_change.destroy()
                        messagebox.showinfo("Notification","Password Changed Successfully!!!")
                        root.deiconify()
                    
                edit_label = ctk.CTkLabel(forgot_change,
                                          text='Input New Password',
                                          font=('arial',30,'underline'))
                edit_label.pack(pady=10)

                pas_label = ctk.CTkLabel(forgot_change,
                                        text="New Password :",
                                        font=('arial',20))
                pas_ntr = ctk.CTkEntry(forgot_change,
                                       width=200,
                                       show='•',
                                       font=('arial',20),
                                       placeholder_text='•PASSWORD')
                
                pas_label.pack(pady=10)
                pas_ntr.pack(pady=10)

                updatebtn = ctk.CTkButton(forgot_change,
                                          width=15,
                                          text='Change Password',
                                          command=lambda:update(),
                                          font=('arial',30),
                                          corner_radius=10,
                                          fg_color='#323232',
                                          hover_color='black',
                                          height=2)
                updatebtn.pack(pady=10)

                forgot_change.mainloop()
            
            else:
                messagebox.showerror('NOTIFICATION','WRONG USERNAME/EMAIL \n TRY AGAIN!')

#=================================================LABEL,ENTRY,BUTTON======================================================================
    forgot_name = ctk.CTkLabel(forgot,
                               text='InPut Your Name :', 
                               width=20,
                               font=('arial',20,'underline'))
    forgot_entry_name = ctk.CTkEntry(forgot, 
                                     width=200,
                                     font=('arial',20),
                                     placeholder_text='•USERNAME')

    forgot_name.pack(pady=(10,5))
    forgot_entry_name.pack(pady=5)

    forgot_email = ctk.CTkLabel(forgot,
                                text='InPut Your Email :', 
                                width=20,
                                font=('arial',20,'underline'))
    forgot_entry = ctk.CTkEntry(forgot, 
                                width=200,
                                font=('arial',20),
                                placeholder_text='•EMAIL')

    forgot_email.pack(pady=5)
    forgot_entry.pack(pady=5)

    enter_btn = ctk.CTkButton(forgot,
                              text="Enter",
                              width=14,
                              command=lambda:go_to_change(),
                              font=('arial',30),
                              fg_color="#323232",
                              hover_color='black')
    enter_btn.pack(pady=5)

#=================================================WHENEVER CLOSES IT GOES BACK TO LOGIN INTERFACE======================================================================
    forgot.protocol("WM_DELETE_WINDOW", lambda: go_back_to_main(root, forgot))
    forgot.mainloop()

#=================================================USER \ PASS ENTRY ======================================================================
user_frame =ctk.CTkFrame(root,width=450,height=50,fg_color='#323232')
user_frame.place(x=80,y=250)

pass_frame =ctk.CTkFrame(root,width=450,height=50,fg_color='#323232')
pass_frame.place(x=80,y=300)

username = ctk.CTkEntry(user_frame,
                 width=400,
                 font=('arial',30),
                 placeholder_text='Username',
                 corner_radius=0,)
password = ctk.CTkEntry(pass_frame,
                 width=400,
                 show="•",
                 font=('arial',30),
                 placeholder_text='Password',
                 corner_radius=0)
login = ctk.CTkButton(root,
               text="LOGIN",
               width=450,
               command=lambda:loginfunction(),
               font=('arial',30),
               anchor='CENTER',
               fg_color='#323232',
               text_color='white',
               hover_color='black')
cancel = ctk.CTkButton(root,
                text="EXIT",
                width=20,
                command=lambda:cancelfunction(),
                font=('arial',30),
                anchor='CENTER',
               fg_color='#323232',
               text_color='white',
               hover_color='black')
register = ctk.CTkButton(root,
                  text="Register",
                  command=lambda:registerfunction(),
                  width=16,
                  font=('arial',25,'underline'),
                  fg_color='#323232',
                hover_color='black')
change_pass = ctk.CTkButton(root,
                            text="Change Password",
                            width=14,
                            command=lambda:change_password(),
                            font=('arial',15,'underline'),
                            fg_color='#323232',
                            hover_color='black')
forgot_password_pass = ctk.CTkButton(root,
                            text="Forgot Password",
                            width=14,
                            command=lambda:forgot_password(),
                            font=('arial',15,'underline'),
                            fg_color='#323232',
                            hover_color='black')
brand_name = ctk.CTkLabel(root,
                        text='ArmorerΩ',
                        font=('courier',30,'underline'),
                        text_color='white',fg_color='black')
brand_label = ctk.CTkLabel(root,
                           text='Equip With Our \nLatest Gun Innovation',
                           font=('courier',25,'bold'),fg_color='black',corner_radius=10)
register_label = ctk.CTkLabel(root,
                              text="DON'T HAVE AN ACCOUNT?",
                              font=('arial',27,'bold'))

username.grid(row=0,column=1)
password.grid(row=0,column=1) 
login.place(x=80,
            y=370)
cancel.place(x=260,
             y=550)
register.place(x=460,
               y=438)
change_pass.place(x=150,
                  y=510)
forgot_password_pass.place(x=300,
                  y=510)
brand_name.place(x=280
                 ,y=70)
brand_label.place(x=160,
                  y=150)
register_label.place(x=80,y=440)

root.mainloop()